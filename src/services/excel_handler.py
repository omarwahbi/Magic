"""Excel file operations service - single responsibility: read/write Excel files."""

import logging
from typing import Set, Dict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.config.settings import AppSettings
from src.config.extraction_config import ExtractionType, ExtractionConfig


class ExcelHandler:
    """Handles all Excel file operations."""

    def __init__(self, file_path: str):
        """
        Initialize Excel handler.

        Args:
            file_path: Path to Excel file
        """
        self.file_path = file_path
        self.settings = AppSettings()

    def read_codes(self) -> Set[str]:
        """
        Read all national codes from Excel file.

        Returns:
            Set of codes found in Excel
        """
        wb = load_workbook(self.file_path)
        ws = wb.active

        codes = set()
        for row in ws.iter_rows(min_row=self.settings.EXCEL_START_ROW):
            code = row[self.settings.CODE_COLUMN].value
            if code and isinstance(code, str):
                codes.add(code.strip().upper())

        wb.close()
        return codes

    def read_code_rows(self) -> Dict[str, int]:
        """
        Read codes and their row numbers.

        Returns:
            Dictionary mapping code -> row number
        """
        wb = load_workbook(self.file_path)
        ws = wb.active

        code_rows = {}
        for row in ws.iter_rows(min_row=self.settings.EXCEL_START_ROW):
            code = row[self.settings.CODE_COLUMN].value
            if code and isinstance(code, str):
                code_upper = code.strip().upper()
                code_rows[code_upper] = row[0].row

        wb.close()
        return code_rows

    def update_balances(
        self,
        balances: Dict[str, float],
        extraction_type: ExtractionType
    ) -> str:
        """
        Update Excel file with extracted balances.

        Args:
            balances: Dictionary of code -> balance
            extraction_type: Type of extraction (determines target column)

        Returns:
            Path to saved file

        Raises:
            Exception: If save fails
        """
        wb = load_workbook(self.file_path)
        ws = wb.active

        # Get target column based on extraction type
        col_idx = ExtractionConfig.get_excel_column(extraction_type)

        updated = 0
        for row in ws.iter_rows(min_row=self.settings.EXCEL_START_ROW):
            code = row[self.settings.CODE_COLUMN].value
            if not code or not isinstance(code, str):
                continue

            code = code.strip().upper()

            if code in balances:
                # Write to target column (convert 0-based to 1-based)
                ws.cell(row=row[0].row, column=col_idx + 1, value=balances[code])
                updated += 1
                logging.debug(f"Updated {code} with balance {balances[code]}")

        # Generate output filename
        timestamp = datetime.now().strftime(self.settings.OUTPUT_DATE_FORMAT)
        output_file = f"{self.settings.OUTPUT_FILE_PREFIX}_{timestamp}.xlsx"

        # Save file
        wb.save(output_file)
        wb.close()

        logging.info(f"Saved {output_file} with {updated} updates")
        return output_file

    def get_worksheet(self) -> Worksheet:
        """
        Get active worksheet (for read-only operations).

        Returns:
            Active worksheet

        Note:
            Caller is responsible for closing the workbook
        """
        wb = load_workbook(self.file_path)
        return wb.active
