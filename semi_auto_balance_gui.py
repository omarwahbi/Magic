import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import pandas as pd
import re
import os
from datetime import datetime
from openpyxl import load_workbook
import logging

# Setup logging
logging.basicConfig(
    filename='extraction_log.txt',
    level=logging.WARNING,
    filemode='w',  # Overwrite log file each time
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class BalanceUpdaterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Semi-Automated Balance Updater")
        self.root.geometry("900x700") # Increased height for new tabs
        
        self.pdf_files = []
        self.excel_file = None
        self.extracted_data = {}
        self.expired_items = []
        self.duplicates = []
        self.extraction_type = tk.StringVar(value="Stock") # Default value
        
        self.setup_ui()
    
    def setup_ui(self):
        # Top frame for file selection
        top_frame = tk.Frame(self.root, padx=10, pady=10)
        top_frame.pack(fill="x")
        
        # PDF selection
        tk.Label(top_frame, text="PDFs:").grid(row=0, column=0, sticky="w")
        self.pdf_label = tk.Label(top_frame, text="No PDFs selected", fg="gray")
        self.pdf_label.grid(row=0, column=1, sticky="w", padx=5)
        tk.Button(top_frame, text="Browse", command=self.select_pdf, bg="#4CAF50", fg="white").grid(row=0, column=2, padx=5)
        
        # Excel selection
        tk.Label(top_frame, text="Excel:").grid(row=1, column=0, sticky="w")
        self.excel_label = tk.Label(top_frame, text="No Excel selected", fg="gray")
        self.excel_label.grid(row=1, column=1, sticky="w", padx=5)
        tk.Button(top_frame, text="Browse", command=self.select_excel, bg="#2196F3", fg="white").grid(row=1, column=2, padx=5)

        # Type selection
        type_frame = tk.Frame(top_frame)
        type_frame.grid(row=2, column=0, columnspan=3, pady=5)
        tk.Label(type_frame, text="Type:").pack(side="left")
        tk.Radiobutton(type_frame, text="Stock", variable=self.extraction_type, value="Stock").pack(side="left", padx=5)
        tk.Radiobutton(type_frame, text="Free", variable=self.extraction_type, value="Free").pack(side="left", padx=5)
        tk.Radiobutton(type_frame, text="Buy", variable=self.extraction_type, value="Buy").pack(side="left", padx=5)
        
        # Extract button
        tk.Button(top_frame, text="Extract from PDF(s)", command=self.extract_data, 
                 bg="#FF9800", fg="white", font=("Arial", 10, "bold")).grid(row=3, column=1, pady=10)
        
        # Main content notebook
        main_notebook = ttk.Notebook(self.root)
        main_notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # --- Results Tabs ---
        results_frame = tk.Frame(main_notebook)
        main_notebook.add(results_frame, text="Extraction Results")
        results_notebook = ttk.Notebook(results_frame)
        results_notebook.pack(fill="both", expand=True)

        # Tab 1: Matched items
        matched_frame = tk.Frame(results_notebook)
        results_notebook.add(matched_frame, text="Matched Items")
        
        columns = ("Code", "Extracted Balance", "Status", "Manual Balance")
        self.tree = ttk.Treeview(matched_frame, columns=columns, show="headings", height=15)
        self.tree.heading("Code", text="National Code")
        self.tree.heading("Extracted Balance", text="Auto-Extracted")
        self.tree.heading("Status", text="Status")
        self.tree.heading("Manual Balance", text="Manual Entry")
        self.tree.column("Code", width=150)
        self.tree.column("Extracted Balance", width=150)
        self.tree.column("Status", width=100)
        self.tree.column("Manual Balance", width=150)
        scrollbar1 = ttk.Scrollbar(matched_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar1.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar1.pack(side="right", fill="y")
        
        # Tab 2: Unmatched codes (in PDF but not in Excel)
        unmatched_frame = tk.Frame(results_notebook)
        results_notebook.add(unmatched_frame, text="Codes NOT in Excel")
        tk.Label(unmatched_frame, text="These codes were found in PDF but don't exist in your Excel file:", font=("Arial", 10, "bold"), fg="red").pack(pady=5)
        columns2 = ("Code", "Balance")
        self.unmatched_tree = ttk.Treeview(unmatched_frame, columns=columns2, show="headings", height=15)
        self.unmatched_tree.heading("Code", text="National Code")
        self.unmatched_tree.heading("Balance", text="Extracted Balance")
        self.unmatched_tree.column("Code", width=200)
        self.unmatched_tree.column("Balance", width=200)
        scrollbar2 = ttk.Scrollbar(unmatched_frame, orient="vertical", command=self.unmatched_tree.yview)
        self.unmatched_tree.configure(yscrollcommand=scrollbar2.set)
        self.unmatched_tree.pack(side="left", fill="both", expand=True)
        scrollbar2.pack(side="right", fill="y")
        tk.Button(unmatched_frame, text="Export to Excel", command=self.export_unmatched_codes, bg="#2196F3", fg="white").pack(pady=5)

        # --- Issues Tabs ---
        issues_frame = tk.Frame(main_notebook)
        main_notebook.add(issues_frame, text="Data Issues")
        issues_notebook = ttk.Notebook(issues_frame)
        issues_notebook.pack(fill="both", expand=True)

        # Tab 3: Expired Items
        expired_frame = tk.Frame(issues_notebook)
        issues_notebook.add(expired_frame, text="Expired Items")
        tk.Label(expired_frame, text="These items were found in the PDF but were skipped because they are expired:", font=("Arial", 10, "bold"), fg="orange").pack(pady=5)
        columns3 = ("Code", "Name", "Expiry Date")
        self.expired_tree = ttk.Treeview(expired_frame, columns=columns3, show="headings", height=15)
        self.expired_tree.heading("Code", text="National Code")
        self.expired_tree.heading("Name", text="Item Name")
        self.expired_tree.heading("Expiry Date", text="Expiry Date")
        self.expired_tree.column("Code", width=150)
        self.expired_tree.column("Name", width=300)
        self.expired_tree.column("Expiry Date", width=100)
        scrollbar3 = ttk.Scrollbar(expired_frame, orient="vertical", command=self.expired_tree.yview)
        self.expired_tree.configure(yscrollcommand=scrollbar3.set)
        self.expired_tree.pack(side="left", fill="both", expand=True)
        scrollbar3.pack(side="right", fill="y")
        tk.Button(expired_frame, text="Export to Excel", command=self.export_expired_items, bg="#2196F3", fg="white").pack(pady=5)

        # Tab 4: Duplicate Codes
        duplicate_frame = tk.Frame(issues_notebook)
        issues_notebook.add(duplicate_frame, text="Duplicate Codes")
        tk.Label(duplicate_frame, text="These national codes are assigned to more than one item in the PDF:", font=("Arial", 10, "bold"), fg="purple").pack(pady=5)
        columns4 = ("Code", "Item Name 1", "Item Name 2")
        self.duplicate_tree = ttk.Treeview(duplicate_frame, columns=columns4, show="headings", height=15)
        self.duplicate_tree.heading("Code", text="National Code")
        self.duplicate_tree.heading("Item Name 1", text="First Item Name")
        self.duplicate_tree.heading("Item Name 2", text="Second Item Name")
        self.duplicate_tree.column("Code", width=150)
        self.duplicate_tree.column("Item Name 1", width=300)
        self.duplicate_tree.column("Item Name 2", width=300)
        scrollbar4 = ttk.Scrollbar(duplicate_frame, orient="vertical", command=self.duplicate_tree.yview)
        self.duplicate_tree.configure(yscrollcommand=scrollbar4.set)
        self.duplicate_tree.pack(side="left", fill="both", expand=True)
        scrollbar4.pack(side="right", fill="y")
        tk.Button(duplicate_frame, text="Export to Excel", command=self.export_duplicates, bg="#2196F3", fg="white").pack(pady=5)
        
        # Manual entry frame
        entry_frame = tk.Frame(self.root, padx=10, pady=10)
        entry_frame.pack(fill="x")
        
        tk.Label(entry_frame, text="Manual correction:").pack(side="left", padx=5)
        tk.Label(entry_frame, text="Balance:").pack(side="left")
        self.manual_entry = tk.Entry(entry_frame, width=15)
        self.manual_entry.pack(side="left", padx=5)
        tk.Button(entry_frame, text="Update Selected", command=self.update_manual, bg="#4CAF50", fg="white").pack(side="left", padx=5)
        
        # Save button
        tk.Button(self.root, text="Save Updated Excel", command=self.save_excel, 
                 bg="#2196F3", fg="white", font=("Arial", 12, "bold"), height=2).pack(pady=10)
        
        # Status
        self.status_label = tk.Label(self.root, text="", fg="blue")
        self.status_label.pack()
    
    def select_pdf(self):
        files = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        if files:
            self.pdf_files = files
            self.pdf_label.config(text=f"{len(files)} PDFs selected", fg="black")
    
    def select_excel(self):
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file:
            self.excel_file = file
            self.excel_label.config(text=os.path.basename(file), fg="black")
    
    def fix_doubled_chars(self, text):
        if not text:
            return text
        return ''.join([text[i] for i in range(0, len(text), 2)])

    def extract_data(self):
        logging.debug("--- Starting Extraction ---")
        if not self.pdf_files or not self.excel_file:
            messagebox.showerror("Error", "Please select PDF(s) and Excel file")
            logging.error("Execution stopped: PDF or Excel file not selected.")
            return

        self.status_label.config(text="Extracting...")
        self.root.update()

        # Show loading screen
        loading_window = tk.Toplevel(self.root)
        loading_window.title("Loading...")
        loading_window.geometry("300x150") # Increased height for progressbar
        loading_window.transient(self.root) # Make it appear on top of the main window
        loading_window.grab_set() # Make it modal
        # Center the loading window
        self.root.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (loading_window.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (loading_window.winfo_height() // 2)
        loading_window.geometry(f"300x150+{x}+{y}")

        tk.Label(loading_window, text="Extracting data... Please wait.", font=("Arial", 12)).pack(pady=20)
        
        progressbar = ttk.Progressbar(loading_window, mode='indeterminate', length=200)
        progressbar.pack()
        progressbar.start()

        aggregated_data = {}
        extraction_summary = []
        extraction_type = self.extraction_type.get()
        logging.debug(f"Extraction type selected: {extraction_type}")

        try:
                for i, pdf_file in enumerate(self.pdf_files):
                    logging.debug(f"Processing PDF {i+1}/{len(self.pdf_files)}: {os.path.basename(pdf_file)}")
                    self.status_label.config(text=f"Processing PDF {i+1}/{len(self.pdf_files)}: {os.path.basename(pdf_file)}...")
                    self.root.update()
                    loading_window.update() # Update loading window to show progressbar

                    current_file_data = {}
                    code_to_name = {}

                    with pdfplumber.open(pdf_file) as pdf:
                        for page_num, page in enumerate(pdf.pages):
                            logging.debug(f"-- Processing Page {page_num + 1} --")
                            tables = page.extract_tables()
                            
                            if not tables:
                                logging.warning(f"No tables found on page {page_num + 1}")
                                continue

                            for table_num, table in enumerate(tables):
                                logging.debug(f"- Processing Table {table_num + 1} on Page {page_num + 1} -")
                                if page_num == 0: # Log table content only for the first page
                                    logging.debug(f"Raw Table Content (Page {page_num + 1}, Table {table_num + 1}): {table}")
                            code_positions = []
                            for r_idx, row in enumerate(table):
                                if not row: continue
                                full_row_text = "".join(str(cell) for cell in row if cell)
                                text_no_spaces = re.sub(r'\\s+', '', full_row_text)
                                text_std_dashes = text_no_spaces.replace('–', '-').replace('—', '-')
                                cleaned_text = self.fix_doubled_chars(text_std_dashes)
                                code_match = re.search(r'([A-Z0-9]{2}-[A-Z0-9]{3}-+[A-Z0-9]{3})', cleaned_text, re.IGNORECASE)
                                if code_match:
                                    code = code_match.group(1).upper()
                                    name = self.fix_doubled_chars(full_row_text.split(code_match.group(0))[0].strip())
                                    logging.debug(f"Found potential code: {code} in table row {r_idx}")
                                    code_positions.append((r_idx, code, name))

                                    # Check for duplicates
                                    if code in code_to_name and code_to_name[code] != name:
                                        self.duplicates.append((code, code_to_name[code], name))
                                    else:
                                        code_to_name[code] = name

                            now = datetime.now()

                            for idx, (row_idx, code, name) in enumerate(code_positions):
                                start_row = row_idx
                                end_row = code_positions[idx + 1][0] if idx + 1 < len(code_positions) else len(table)
                                logging.debug(f"Processing code '{code}'. Item row range in table: {start_row} to {end_row - 1}")

                                is_expired = False
                                expiry_date_str = ""
                                for r in range(start_row, end_row):
                                    row_text = "".join(str(cell) for cell in table[r] if cell)
                                    cleaned_row_text = self.fix_doubled_chars(row_text)
                                    logging.debug(f"      Scanning row for expiry: '{cleaned_row_text}'")
                                    match = re.search(r'(\\d{1,2})/(\\d{1,2})/(\\d{4})', cleaned_row_text)
                                    if match:
                                        logging.debug(f"        Expiry regex matched: {match.groups()}")
                                        try:
                                            day = int(match.group(1))
                                            month = int(match.group(2))
                                            year = int(match.group(3))
                                            expiry_date_str = f"{day}/{month}/{year}"
                                            if year < now.year or (year == now.year and month < now.month):
                                                is_expired = True
                                                logging.debug(f"Code {code} is EXPIRED with date {expiry_date_str}. Skipping.")
                                                self.expired_items.append((code, name, expiry_date_str))
                                            break
                                        except ValueError:
                                            continue
                                
                                if is_expired:
                                    continue

                                # Simplified logic for all types
                                balance_row_idx = end_row - 1
                                while balance_row_idx > row_idx:
                                    if any(table[balance_row_idx]): break
                                    balance_row_idx -= 1
                                
                                if balance_row_idx < len(table):
                                    balance_row = table[balance_row_idx]
                                    logging.debug(f"      >>> Found balance row! Content: {balance_row}")
                                    # Conditional column index based on extraction type
                                    if extraction_type == "Free":
                                        col_idx = 2 # الوارد (Incoming) - from اورام الشركة العامة اسكان.pdf
                                    elif extraction_type == "Stock":
                                        col_idx = 7 # Actual Balance - from المساحيق.pdf
                                    else: # Buy
                                        col_idx = 2 # الوارد (Incoming) - from مضادات شركة عامة.pdf
                                    
                                    if len(balance_row) > col_idx:
                                        cell = balance_row[col_idx]
                                        if cell:
                                            try:
                                                balance_str = str(cell).replace(',', '').strip()
                                                balance = float(balance_str)
                                                logging.debug(f"        >>> Found balance for '{code}': {balance}")
                                                if code in current_file_data: current_file_data[code] += balance
                                                else: current_file_data[code] = balance
                                            except (ValueError, TypeError):
                                                logging.warning(f"        Could not convert cell '{cell}' to a number.")
                                    else:
                                        logging.warning(f"      Balance row does not have the required column index: {col_idx}")
        except Exception as e:
            logging.error(f"Error during PDF extraction: {e}")
            messagebox.showerror("Error", f"An error occurred during PDF extraction: {e}")
        finally:
            progressbar.stop()
            loading_window.destroy()
            self.status_label.config(text="Extraction complete.")
            self.display_results() # Call display_results after extraction

    def display_results(self):
        # Clear all trees
        for tree in [self.tree, self.unmatched_tree, self.expired_tree, self.duplicate_tree]:
            for item in tree.get_children():
                tree.delete(item)
        
        # Load Excel to get all codes
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        excel_codes = set()
        
        # First pass: populate matched items
        for row in ws.iter_rows(min_row=2):
            code = row[0].value
            if not code or not isinstance(code, str):
                continue
            
            code = str(code).strip().upper()
            excel_codes.add(code)
            
            if code in self.extracted_data:
                balance = self.extracted_data[code]
                status = "✓ OK"
                tag = "success"
            else:
                balance = ""
                status = "✗ Missing"
                tag = "missing"
            
            self.tree.insert("", "end", values=(code, balance, status, ""), tags=(tag,))
        
        # Second pass: find codes in PDF but not in Excel
        unmatched_codes = []
        for code, balance in self.extracted_data.items():
            if code not in excel_codes:
                unmatched_codes.append((code, balance))
                self.unmatched_tree.insert("", "end", values=(code, balance))
        
        # Populate expired items tab
        for code, name, expiry_date in self.expired_items:
            self.expired_tree.insert("", "end", values=(code, name, expiry_date))

        # Populate duplicate codes tab
        for code, name1, name2 in self.duplicates:
            self.duplicate_tree.insert("", "end", values=(code, name1, name2))

        # Color coding
        self.tree.tag_configure("success", background="#d4edda")
        self.tree.tag_configure("missing", background="#f8d7da")
        
        # Update status
        matched_count = sum(1 for code in excel_codes if code in self.extracted_data)
        missing_count = len(excel_codes) - matched_count
        
        self.status_label.config(
            text=f"Matched: {matched_count} | Missing: {missing_count} | Not in Excel: {len(unmatched_codes)} | Duplicates: {len(self.duplicates)} | Expired: {len(self.expired_items)}"
        )
    
    def update_manual(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a row first")
            return
        
        manual_value = self.manual_entry.get().strip()
        if not manual_value:
            messagebox.showwarning("Warning", "Please enter a balance value")
            return
        
        try:
            balance = float(manual_value)
            item = selected[0]
            code = self.tree.item(item)['values'][0]
            
            # Update extracted data
            self.extracted_data[code] = balance
            
            # Update display
            self.tree.item(item, values=(code, balance, "✓ Manual", ""), tags=("success",))
            self.manual_entry.delete(0, tk.END)
            
        except ValueError:
            messagebox.showerror("Error", "Invalid number format")
    
    def save_excel(self):
        if not self.extracted_data:
            messagebox.showwarning("Warning", "No data to save")
            return
        
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Determine target column based on type
            extraction_type = self.extraction_type.get()
            if extraction_type == "Free":
                col_idx = 5 # Column F
            elif extraction_type == "Buy":
                col_idx = 7 # Column H
            else: # Stock
                col_idx = 6 # Column G

            updated = 0
            for row in ws.iter_rows(min_row=2):
                code = row[0].value
                if not code or not isinstance(code, str):
                    continue
                
                code = str(code).strip().upper()
                
                if code in self.extracted_data:
                    # Write to the correct column (1-based index for user, so col_idx+1)
                    ws.cell(row=row[0].row, column=col_idx + 1, value=self.extracted_data[code])
                    updated += 1
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"updated_medicines_{timestamp}.xlsx"
            wb.save(output_file)
            
            messagebox.showinfo("Success", 
                              f"Saved successfully!\n\n"
                              f"Updated: {updated} items\n"
                              f"File: {output_file}")
            
            self.status_label.config(text=f"✓ Saved: {output_file}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Save failed: {str(e)}")

    def export_unmatched_codes(self):
        if not self.unmatched_tree.get_children():
            messagebox.showwarning("Warning", "No unmatched codes to export.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            data = []
            for item_id in self.unmatched_tree.get_children():
                data.append(self.unmatched_tree.item(item_id)['values'])
            
            df = pd.DataFrame(data, columns=["Code", "Balance"])
            df.to_excel(file_path, index=False)
            messagebox.showinfo("Success", f"Unmatched codes exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export unmatched codes: {str(e)}")

    def export_expired_items(self):
        if not self.expired_tree.get_children():
            messagebox.showwarning("Warning", "No expired items to export.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            data = []
            for item_id in self.expired_tree.get_children():
                data.append(self.expired_tree.item(item_id)['values'])
            
            df = pd.DataFrame(data, columns=["Code", "Name", "Expiry Date"])
            df.to_excel(file_path, index=False)
            messagebox.showinfo("Success", f"Expired items exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export expired items: {str(e)}")

    def export_duplicates(self):
        if not self.duplicate_tree.get_children():
            messagebox.showwarning("Warning", "No duplicate codes to export.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            data = []
            for item_id in self.duplicate_tree.get_children():
                data.append(self.duplicate_tree.item(item_id)['values'])
            
            df = pd.DataFrame(data, columns=["Code", "Item Name 1", "Item Name 2"])
            df.to_excel(file_path, index=False)
            messagebox.showinfo("Success", f"Duplicate codes exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export duplicate codes: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = BalanceUpdaterGUI(root)
    root.mainloop()
